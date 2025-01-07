from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, Q
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
from .models import Graduate, GraduateBulkUpload, ExamCenter, Course, District
from .forms import GraduateForm, GraduateBulkUploadForm
import logging
import traceback
from django.core.exceptions import ObjectDoesNotExist

logger = logging.getLogger(__name__)

def is_admin(user):
    return user.user_type == 'admin'

def is_field_officer(user):
    return user.user_type == 'field_officer'

def is_officer(user):
    return user.user_type == 'officer'

def can_manage_data(user):
    """Users who can add/edit graduate data"""
    return user.user_type in ['admin', 'field_officer']

def can_view_reports(user):
    """Users who can view and generate reports"""
    return user.user_type in ['admin', 'officer']

def can_manage_users(user):
    """Only admins can manage user accounts"""
    return user.user_type == 'admin'

@login_required
def dashboard(request):
    try:
        logger.info(f"Dashboard accessed by user: {request.user.username}")
        
        # Print request information
        logger.info(f"Request method: {request.method}")
        logger.info(f"Request path: {request.path}")
        logger.info(f"User type: {request.user.user_type}")
        
        # Basic statistics
        try:
            total_graduates = Graduate.objects.count()
            logger.info(f"Total graduates: {total_graduates}")
        except Exception as e:
            logger.error(f"Error getting total graduates: {str(e)}")
            total_graduates = 0
            
        try:
            employed_graduates = Graduate.objects.filter(is_employed=True).count()
            logger.info(f"Employed graduates: {employed_graduates}")
        except Exception as e:
            logger.error(f"Error getting employed graduates: {str(e)}")
            employed_graduates = 0
        
        # Graduates by year
        try:
            graduates_by_year = list(Graduate.objects.values('graduation_date__year')
                                   .annotate(count=Count('id'))
                                   .order_by('graduation_date__year'))
            logger.info(f"Graduates by year: {graduates_by_year}")
        except Exception as e:
            logger.error(f"Error getting graduates by year: {str(e)}")
            graduates_by_year = []
        
        # Employment statistics by course
        try:
            course_stats = list(Graduate.objects.values('course__name')
                               .annotate(
                                   total=Count('id'),
                                   employed=Count('id', filter=Q(is_employed=True))
                               )
                               .order_by('-total'))
            logger.info(f"Course stats: {course_stats}")
        except Exception as e:
            logger.error(f"Error getting course stats: {str(e)}")
            course_stats = []
        
        # Gender distribution
        try:
            gender_stats = list(Graduate.objects.values('gender')
                               .annotate(count=Count('id'))
                               .order_by('gender'))
            logger.info(f"Gender stats: {gender_stats}")
        except Exception as e:
            logger.error(f"Error getting gender stats: {str(e)}")
            gender_stats = []
        
        # Recent graduates
        try:
            if request.user.user_type == 'field_officer':
                recent_graduates = Graduate.objects.filter(created_by=request.user).order_by('-created_at')[:5]
            else:
                recent_graduates = Graduate.objects.all().order_by('-created_at')[:5]
            logger.info(f"Recent graduates count: {len(recent_graduates)}")
        except Exception as e:
            logger.error(f"Error getting recent graduates: {str(e)}")
            recent_graduates = []
        
        # Calculate employment rate safely
        employment_rate = (employed_graduates / total_graduates * 100) if total_graduates > 0 else 0
        
        context = {
            'total_graduates': total_graduates,
            'employed_graduates': employed_graduates,
            'employment_rate': employment_rate,
            'graduates_by_year': graduates_by_year,
            'course_stats': course_stats,
            'gender_stats': gender_stats,
            'recent_graduates': recent_graduates,
            'user_type': request.user.user_type,
        }
        
        logger.info("Dashboard context prepared successfully")
        return render(request, 'graduates/dashboard.html', context)
        
    except Exception as e:
        logger.error(f"Dashboard error: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        # Return a simple error page with basic stats
        context = {
            'error': True,
            'error_message': str(e),
            'total_graduates': Graduate.objects.count(),
            'user_type': request.user.user_type,
        }
        return render(request, 'graduates/dashboard.html', context)

@login_required
@user_passes_test(can_manage_data)
def add_graduate(request):
    if request.method == 'POST':
        form = GraduateForm(request.POST)
        if form.is_valid():
            graduate = form.save(commit=False)
            graduate.created_by = request.user
            graduate.save()
            messages.success(request, 'Graduate added successfully!')
            return redirect('dashboard')
    else:
        form = GraduateForm()
    
    return render(request, 'graduates/add_graduate.html', {'form': form})

@login_required
def graduate_list(request):
    if request.user.user_type == 'field_officer':
        # Field officers only see their own submissions
        graduates = Graduate.objects.filter(created_by=request.user).order_by('-created_at')
    else:
        # Admins and officers see all
        graduates = Graduate.objects.all().order_by('-created_at')
    
    return render(request, 'graduates/graduate_list.html', {
        'graduates': graduates,
        'can_manage_data': can_manage_data(request.user),
        'can_view_reports': can_view_reports(request.user),
    })

@login_required
@user_passes_test(can_manage_data)
def upload_graduates(request):
    if request.method == 'POST':
        form = GraduateBulkUploadForm(request.POST, request.FILES)
        if form.is_valid():
            upload = form.save(commit=False)
            upload.uploaded_by = request.user
            upload.save()
            
            try:
                df = pd.read_excel(upload.file.path)
                
                # Debug: Print column names and first row
                print("Columns in Excel:", df.columns.tolist())
                if not df.empty:
                    print("First row:", df.iloc[0].to_dict())
                
                required_columns = [
                    'first_name', 'last_name', 'gender', 'date_of_birth',
                    'contact_number', 'email', 'registration_number',
                    'course_name', 'graduation_year', 'is_employed',
                    'current_district', 'exam_center'
                ]
                
                # Convert column names to lowercase and strip whitespace
                df.columns = df.columns.str.lower().str.strip()
                
                # Debug: Print columns after conversion
                print("Columns after conversion:", df.columns.tolist())
                
                # Verify required columns
                missing_columns = [col for col in required_columns if col not in df.columns]
                if missing_columns:
                    raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")
                
                # Process each row
                successful_imports = 0
                for index, row in df.iterrows():
                    try:
                        # Convert 'Yes'/'No' to boolean for is_employed
                        is_employed = str(row['is_employed']).lower() in ['yes', 'true', '1', 'y']
                        
                        # Convert date fields
                        try:
                            date_of_birth = pd.to_datetime(row['date_of_birth']).date()
                        except:
                            date_of_birth = None
                            print(f"Invalid date_of_birth format in row {index + 2}")
                        
                        try:
                            employment_date = pd.to_datetime(row['employment_date']).date() if pd.notna(row.get('employment_date')) else None
                        except:
                            employment_date = None
                            print(f"Invalid employment_date format in row {index + 2}")
                        
                        # Get or create course
                        course, _ = Course.objects.get_or_create(
                            name=str(row['course_name']),
                            defaults={'department': 'Default Department'}
                        )
                        
                        # Get district and exam center
                        district_name = str(row.get('current_district', '')).strip()
                        try:
                            district = District.objects.get(name__iexact=district_name)
                        except District.DoesNotExist:
                            messages.warning(request, f'District "{district_name}" in row {index + 2} not found. Please choose from the provided list.')
                            continue
                            
                        try:
                            exam_center = ExamCenter.objects.get(name=str(row['exam_center']))
                        except ExamCenter.DoesNotExist:
                            exam_center = None
                            print(f"Exam center not found in row {index + 2}")
                        
                        Graduate.objects.create(
                            first_name=str(row['first_name']),
                            last_name=str(row['last_name']),
                            gender=str(row['gender']),
                            date_of_birth=date_of_birth,
                            contact_number=str(row['contact_number']),
                            email=str(row['email']),
                            registration_number=str(row['registration_number']),
                            course=course,
                            graduation_year=int(float(row['graduation_year'])),
                            is_employed=is_employed,
                            employer_name=str(row.get('employer_name', '')),
                            job_title=str(row.get('job_title', '')),
                            employment_start_date=employment_date,
                            current_district=district,
                            exam_center=exam_center,
                            created_by=request.user
                        )
                        successful_imports += 1
                    except Exception as row_error:
                        print(f"Error in row {index + 2}: {str(row_error)}")
                        continue
                
                upload.is_processed = True
                upload.save()
                
                if successful_imports > 0:
                    messages.success(request, f'Successfully imported {successful_imports} graduates!')
                else:
                    messages.warning(request, 'No graduates were imported. Please check the file format.')
                return redirect('graduate_list')
            except ValueError as e:
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
    else:
        form = GraduateBulkUploadForm()
    
    return render(request, 'graduates/upload_graduates.html', {'form': form})

@login_required
@user_passes_test(can_manage_data)
def download_excel_template(request):
    # Create Excel template
    df = pd.DataFrame(columns=[
        'first_name', 'last_name', 'gender', 'date_of_birth',
        'contact_number', 'email', 'registration_number',
        'course_name', 'graduation_year', 'is_employed',
        'employer_name', 'job_title', 'employment_date',
        'current_district', 'exam_center'
    ])
    
    # Create a response with Excel MIME type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=graduate_upload_template.xlsx'
    
    # Save DataFrame to Excel with formatting
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Template')
        worksheet = writer.sheets['Template']
        
        # Format header row
        for col in worksheet.iter_cols(min_row=1, max_row=1):
            cell = col[0]
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Add data validation for gender
        gender_validation = openpyxl.worksheet.datavalidation.DataValidation(
            type="list",
            formula1='"M,F"',
            allow_blank=True
        )
        worksheet.add_data_validation(gender_validation)
        gender_col = openpyxl.utils.get_column_letter(3)  # Column C for gender
        gender_validation.add(f'{gender_col}2:{gender_col}1000')
        
        # Add data validation for is_employed
        employed_validation = openpyxl.worksheet.datavalidation.DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=True
        )
        worksheet.add_data_validation(employed_validation)
        employed_col = openpyxl.utils.get_column_letter(10)  # Column J for is_employed
        employed_validation.add(f'{employed_col}2:{employed_col}1000')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width
    
    return response

@login_required
@user_passes_test(can_view_reports)
def export_graduates_excel(request):
    try:
        # Get graduates with related fields
        graduates = Graduate.objects.select_related('exam_center', 'exam_center__district', 'course').all()
        
        # Prepare data for Excel
        data = []
        for graduate in graduates:
            data.append({
                'Registration Number': graduate.registration_number,
                'First Name': graduate.first_name,
                'Last Name': graduate.last_name,
                'Gender': graduate.get_gender_display(),
                'Date of Birth': graduate.date_of_birth.strftime('%Y-%m-%d') if graduate.date_of_birth else '',
                'Email': graduate.email,
                'Phone Number': graduate.phone_number,
                'Course': graduate.course.name if graduate.course else '',
                'Graduation Date': graduate.graduation_date.strftime('%Y-%m-%d') if graduate.graduation_date else '',
                'Exam Center': graduate.exam_center.name if graduate.exam_center else '',
                'District': graduate.exam_center.district.name if graduate.exam_center and graduate.exam_center.district else '',
                'Employment Status': 'Employed' if graduate.is_employed else 'Not Employed',
                'Employer': graduate.employer_name if graduate.is_employed else '',
                'Job Title': graduate.job_title if graduate.is_employed else '',
                'Employment Date': graduate.employment_date.strftime('%Y-%m-%d') if graduate.employment_date else '',
            })
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=graduates.xlsx'
        
        # Create Excel writer
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Graduates')
            
            # Get the worksheet
            worksheet = writer.sheets['Graduates']
            
            # Format headers
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                
                # Find the maximum length in the column
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set the column width
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column].width = adjusted_width
                
                # Format header row
                header_cell = worksheet[f"{column}1"]
                header_cell.font = openpyxl.styles.Font(bold=True)
                header_cell.fill = openpyxl.styles.PatternFill(start_color='333333', end_color='333333', fill_type='solid')
                header_cell.font = openpyxl.styles.Font(color='FFFFFF', bold=True)
        
        return response
        
    except Exception as e:
        logger.error(f"Error exporting Excel: {str(e)}", exc_info=True)
        messages.error(request, "An error occurred while generating the Excel file. Please try again.")
        return redirect('graduate_list')

@login_required
@user_passes_test(can_view_reports)
def export_graduates_pdf(request):
    try:
        # Get all graduates
        graduates = Graduate.objects.select_related('exam_center', 'exam_center__district').all()
        
        # Create the HttpResponse object with PDF headers
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=graduates.pdf'
        
        # Generate PDF using our utility function
        from .utils import generate_graduate_pdf
        generate_graduate_pdf(response, graduates)
        
        return response
        
    except Exception as e:
        logger.error(f"Error generating PDF: {str(e)}", exc_info=True)
        messages.error(request, "An error occurred while generating the PDF. Please try again.")
        return redirect('graduate_list')

@login_required
@user_passes_test(can_manage_data)
def edit_graduate(request, pk):
    graduate = get_object_or_404(Graduate, pk=pk)
    
    # Field officers can only edit their own submissions
    if request.user.user_type == 'field_officer' and graduate.created_by != request.user:
        messages.error(request, 'You can only edit your own submissions.')
        return redirect('graduate_list')
    
    if request.method == 'POST':
        form = GraduateForm(request.POST, instance=graduate)
        if form.is_valid():
            form.save()
            messages.success(request, 'Graduate updated successfully!')
            return redirect('graduate_list')
    else:
        form = GraduateForm(instance=graduate)
    
    return render(request, 'graduates/edit_graduate.html', {'form': form})

@login_required
def get_exam_centers(request):
    district_id = request.GET.get('current_district')
    exam_centers = ExamCenter.objects.all().values('id', 'name')
    return JsonResponse(list(exam_centers), safe=False)

@login_required
def get_programs(request):
    programs = Course.objects.all().values('id', 'name')
    return JsonResponse(list(programs), safe=False)

@login_required
def get_districts(request):
    districts = District.objects.all().values('id', 'name').order_by('name')
    return JsonResponse(list(districts), safe=False)
